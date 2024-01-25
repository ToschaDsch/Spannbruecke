import io
import sys
from functools import partial
from tkinter.filedialog import askopenfilename
import numpy as np
import openpyxl
from PySide6 import QtGui, QtWidgets
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QMainWindow, QApplication, QLabel, QVBoxLayout, QTableWidget, QWidget, QTableWidgetItem

from class_cable import Cable
from variables import Variables, Result


class WindowSection(QMainWindow):
    def __init__(self):
        super().__init__()
        self.dict_of_cables = {}
        self._graph_scale_x: float = 1
        self._graph_scale_y: float = 1
        self._graph_x0: float = 10
        self._graph_y0: float = 0
        self._h_of_beam: int = 145  # height in cm
        self._label = QLabel('hello')
        self._table: [QTableWidget] = []
        self._tab_menu = QtWidgets.QTabWidget()
        self._x0_for_tab = 0
        self._x0_dict = {}
        self._dict_for_x = {}
        self._dict_matrix_of_yi = {}
        self._result: dict = {}
        self._sheets = []

        self.canvas = QtGui.QPixmap(Variables.screen_bh[0], Variables.screen_bh[1])
        self.canvas.fill(Qt.black)
        self._label.setPixmap(self.canvas)
        self.painter: QtGui.QPainter | None = None
        self.general_layout = QVBoxLayout()
        self.general_layout.addWidget(self._label)
        self.general_layout.addWidget(self._tab_menu)
        widget = QWidget()
        widget.setLayout(self.general_layout)
        self.setCentralWidget(widget)
        self.open_file()
        self._make_result()
        self._scale_the_graph()
        self._draw_graph()

    def _make_result(self):
        for sheet_name in self._sheets:
            self._result[sheet_name] = []
            for i in range(len(self._dict_for_x[sheet_name])):
                x = self._dict_for_x[sheet_name][i]
                n = 0
                sy = 0
                print(self._dict_matrix_of_yi[sheet_name])
                for y_i in self._dict_matrix_of_yi[sheet_name]:
                    if y_i[i]:
                        n += 1
                        sy += y_i[i]
                if n:
                    m_y = sy/n
                    self._result[sheet_name].append(Result(x=x, y=m_y, n=n))

    def _draw_result(self):
        for sheet, list_of_result in self._result.items():
            self._draw_result_for_a_sheet(name_of_sheet=sheet, list_of_result=list_of_result)

    def _draw_result_for_a_sheet(self, name_of_sheet: str, list_of_result: [Result]):
        for i in range(len(list_of_result) - 1):
            x0i = list_of_result[i].x
            y0i = list_of_result[i].y
            n0 = list_of_result[i].n
            x1i = list_of_result[i + 1].x
            y1i = list_of_result[i + 1].y
            n1 = list_of_result[i + 1].n
            color = (180, 180, 180)
            pen = QtGui.QPen()
            pen.setColor(QtGui.QColor(QColor(*color)))
            pen.setWidth(n0)
            self.painter.setPen(pen)
            x0 = self._x0_dict[name_of_sheet]
            x1 = self._graph_scale_x * (x0 + x0i) + self._graph_x0
            x2 = self._graph_scale_x * (x0 + x1i) + self._graph_x0
            y1 = Variables.screen_bh[1] - self._graph_scale_y * (y0i - self._graph_y0) - 10
            y2 = Variables.screen_bh[1] - self._graph_scale_y * (y1i - self._graph_y0) - 10
            self.painter.drawLine(x1, y1, x2, y2)
            print(x1, y1, x2, y2)



    def _draw_background_and_the_beam(self):
        # draw background
        x = list(self._x0_dict.values())
        print(x, self._x0_for_tab)
        dx = self._graph_scale_x * (x[-1] - x[-2])
        x0 = self._graph_scale_x * x[-2] + self._graph_x0
        x1 = self._graph_scale_x * x[-1] + self._graph_x0
        print(x1, x0)
        y0 = 0
        y1 = Variables.screen_bh[1]
        color = QColor(*(50, 50, 50))
        brush = QtGui.QBrush(color)
        self.painter.setBrush(brush)
        self.painter.drawRect(x0, y0, dx, y1)
        # draw the beam
        pen = QtGui.QPen()
        pen.setColor(QtGui.QColor(QColor(*(150, 150, 150))))
        self.painter.setPen(pen)
        x0 = self._graph_x0
        x1 = Variables.screen_bh[0] - self._graph_x0
        y0 += 10
        y1 -= 10
        self.painter.drawLine(x0, y0, x1, y0)
        self.painter.drawLine(x1, y0, x1, y1)
        self.painter.drawLine(x1, y1, x0, y1)
        self.painter.drawLine(x0, y1, x0, y0)

    def _draw_graph(self):
        canvas = self._label.pixmap()
        canvas.fill(Qt.black)
        self.painter = QtGui.QPainter(canvas)
        pen = QtGui.QPen()
        self._draw_background_and_the_beam()

        for name_of_sheet, group_of_cables in self.dict_of_cables.items():
            for key, cable in group_of_cables.items():
                color = cable.color
                pen.setColor(QtGui.QColor(QColor(*color)))
                if cable.selected:
                    pen.setWidth(3)
                else:
                    pen.setWidth(1)
                self.painter.setPen(pen)
                self._draw_a_cable(cable, name_of_sheet=name_of_sheet)
        self._draw_result()
        self.painter.end()
        self._label.setPixmap(canvas)

    def _draw_a_cable(self, cable: Cable, name_of_sheet: str):
        coordinate: [] = cable.coordinate
        for i in range(len(coordinate) - 1):
            x0 = self._x0_dict[name_of_sheet]
            x1 = self._graph_scale_x * (x0 + coordinate[i][0]) + self._graph_x0
            x2 = self._graph_scale_x * (x0 + coordinate[i + 1][0]) + self._graph_x0
            y1 = Variables.screen_bh[1] - self._graph_scale_y * (coordinate[i][1] - self._graph_y0) - 10
            y2 = Variables.screen_bh[1] - self._graph_scale_y * (coordinate[i + 1][1] - self._graph_y0) - 10
            self.painter.drawLine(x1, y1, x2, y2)
        self._draw_begin(cable=cable, name_of_sheet=name_of_sheet)
        self._draw_end(cable=cable, name_of_sheet=name_of_sheet)

    def _draw_begin(self, cable: Cable, name_of_sheet: str):
        if cable.begin is None:
            return None
        x0 = self._x0_dict[name_of_sheet]
        a = 5
        x0 = self._graph_scale_x * (x0 + cable.begin[0]) + self._graph_x0
        y0 = Variables.screen_bh[1] - self._graph_scale_y * (cable.begin[1] - self._graph_y0) - 10

        if cable.end is None or cable.begin[0] < cable.end[0]:
            x1 = x0 - a
            y1 = y0 - a
            x2 = x1
            y2 = y0 + a
            x3 = x0 + a
            y3 = y0
        else:
            x1 = x0 + a
            y1 = y0 - a
            x2 = x1
            y2 = y0 + a
            x3 = x0 - a
            y3 = y0

        self.painter.drawLine(x1, y1, x2, y2)
        self.painter.drawLine(x3, y3, x2, y2)
        self.painter.drawLine(x3, y3, x1, y1)

    def _draw_end(self, cable: Cable, name_of_sheet: str):
        if cable.end is None:
            return None
        x0 = self._x0_dict[name_of_sheet]
        x1 = self._graph_scale_x * (x0 + cable.end[0]) + self._graph_x0
        y1 = Variables.screen_bh[1] - self._graph_scale_y * (cable.end[1] - self._graph_y0) - 10

        self._draw_an_end_anker(x1, y1)

        if cable.end_2 is None:
            return None
        x1 = self._graph_scale_x * (x0 + cable.end_2[0]) + self._graph_x0
        y1 = Variables.screen_bh[1] - self._graph_scale_y * (cable.end_2[1] - self._graph_y0) - 10
        self._draw_an_end_anker(x1, y1)

    def _draw_an_end_anker(self, x0: float, y0: float):
        a = 5
        x1 = x0 - a
        y1 = y0 - a
        x2 = x0 + a
        y2 = y0 - a
        x3 = x0 + a
        y3 = y0 + a
        x4 = x0 - a
        y4 = y0 + a

        self.painter.drawLine(x1, y1, x2, y2)
        self.painter.drawLine(x3, y3, x2, y2)
        self.painter.drawLine(x3, y3, x4, y4)
        self.painter.drawLine(x1, y1, x4, y4)

    def _scale_the_graph(self):
        if self._x0_for_tab == 0:
            return None
        self._graph_scale_x = (Variables.screen_bh[0] - 2 * self._graph_x0) / self._x0_for_tab
        y_min = 1000
        y_max = -1000
        for name_of_sheet, group_of_cables in self.dict_of_cables.items():
            for key, cable in group_of_cables.items():
                coordinate: [] = cable.coordinate
                for x, y in coordinate:
                    if y > y_max:
                        y_max = y
                    elif y < y_min:
                        y_min = y
        self._graph_y0 = 0
        self._graph_scale_y = (Variables.screen_bh[1] - 20) / self._h_of_beam

    def open_file(self):
        file_name = askopenfilename()
        print("the file is open -")
        print(file_name)

        with (open(file_name, "rb") as f):
            in_mem_file = io.BytesIO(f.read())
            work_book = openpyxl.load_workbook(in_mem_file, read_only=True)
            print(work_book)
            print(work_book.sheetnames)
            self._sheets = work_book.sheetnames
            for i, sheet in enumerate(work_book.sheetnames):
                self.make_data(sheet=work_book.worksheets[i], sheet_name=work_book.sheetnames[i])
                self.add_a_tab_with_data(sheet_name=work_book.sheetnames[i])
            work_book.close()
        print('data is ready')

    def add_a_tab_with_data(self, sheet_name: str):
        self._tab_menu.currentChanged.connect(self._change_tab)
        tab_widget = QWidget()
        tab_layout = QVBoxLayout()
        self._table.append(QTableWidget())
        tab_layout.addWidget(self._table[-1])
        tab_widget.setLayout(tab_layout)
        self._tab_menu.addTab(tab_widget, str(sheet_name))
        self.init_table(sheet_name=sheet_name, table=self._table[-1])

    def _change_tab(self):
        pass

    def new_tab_layout_with_data(self):
        pass

    def make_data(self, sheet, sheet_name):
        dict_of_cable = {}
        set_of_current_cables: set[int] = set()
        first_value_x_direction = 4
        i = first_value_x_direction
        y_for_x_value = 3
        list_x = []
        x: float = 0
        while sheet.cell(row=y_for_x_value, column=i).value is not None:
            x += float(sheet.cell(row=y_for_x_value, column=i).value)
            list_x.append(x)
            i += 1
        self._x0_dict[sheet_name] = self._x0_for_tab
        self._x0_for_tab += list_x[-1]
        self._dict_for_x[sheet_name] = list_x
        n_0y = 0
        while sheet.cell(row=4 + n_0y, column=2).value is not None:
            n_0y += 1

        for y in range(4, 4 + n_0y):
            i = first_value_x_direction
            x = 0
            while sheet.cell(row=y_for_x_value, column=i).value is not None:
                x += float(sheet.cell(row=y_for_x_value, column=i).value)
                value = str(sheet.cell(row=y, column=i).value)
                if value is None or value == 'None':
                    i += 1
                    continue
                if value[-1] == ')':
                    self._end_or_begin_of_the_cables(value=value, set_of_current_cables=set_of_current_cables,
                                                     dict_of_cable=dict_of_cable, x=x, i=i,
                                                     first_value_x_direction=first_value_x_direction,
                                                     sheet_name=sheet_name)
                else:
                    y_i = float(value)
                    for name_of_cable in set_of_current_cables:
                        cable: Cable = dict_of_cable[name_of_cable]
                        cable.coordinate.append([x, y_i])
                i += 1

        self.dict_of_cables[sheet_name] = dict_of_cable

    def _end_or_begin_of_the_cables(self, value: str, set_of_current_cables: set, dict_of_cable: dict,
                                    x: float, i: int, first_value_x_direction: int, sheet_name: str):
        y_i, end_begin, numbers_of_cable = self._make_variables_from_the_string(value=value)
        for number_of_cable in set_of_current_cables:
            cable: Cable = dict_of_cable[number_of_cable]
            cable.coordinate.append([x, y_i])
        for number_of_cable in numbers_of_cable:
            if number_of_cable in dict_of_cable:  # remove an old cable
                cable: Cable = dict_of_cable[number_of_cable]
                cable.i_n = i - first_value_x_direction
                set_of_current_cables.discard(number_of_cable)
            else:  # add a new cable
                cable = Cable(name=number_of_cable, name_of_tab=sheet_name)
                print(f'+ cable {cable.name}')
                cable.color = list(np.random.choice(range(50, 256), size=3))
                dict_of_cable[number_of_cable] = cable
                cable.i_0 = i - first_value_x_direction
                set_of_current_cables.add(number_of_cable)
                cable.coordinate.append([x, y_i])
            if end_begin == Variables.end:
                if cable.end is None:
                    cable.end = [x, y_i]
                    print('cable end', cable.name, cable.end)
                else:
                    cable.end_2 = [x, y_i]
                    print('cable end_2', cable.name, cable.end_2, 'cable end', cable.end)
            elif end_begin == Variables.begin:
                if cable.begin:
                    print('!!!cable begin', cable.name)
                cable.begin = [x, y_i]

    @staticmethod
    def _make_variables_from_the_string(value) -> tuple[float, str, []]:
        n = value.find(' ')
        y_i = value[:n]
        nn = value[n + 1:]
        y_i = y_i.replace(',', '.')
        y_i = float(y_i)
        nn = nn[1: -1]
        if nn.find(Variables.end) != -1:
            end_begin = Variables.end
            nn = nn.replace(Variables.end, '')
        else:
            end_begin = Variables.begin
            nn = nn.replace(Variables.begin, '')
        numbers_of_cable = nn.split(', ')
        numbers_of_cable = [u.replace(')', '') for u in numbers_of_cable]
        numbers_of_cable = [u.replace('(', '') for u in numbers_of_cable]
        numbers_of_cable = [int(x) for x in numbers_of_cable]
        return y_i, end_begin, numbers_of_cable

    def init_table(self, sheet_name: str, table: QTableWidget):
        new_list_of_cable = self.get_a_new_list_of_cable(sheet_name=sheet_name)
        list_for_x = self._dict_for_x[sheet_name]
        table.setFixedWidth(Variables.screen_bh[0])
        table.setColumnCount(len(list_for_x))
        table.setRowCount(len(new_list_of_cable))
        table.itemSelectionChanged.connect(partial(self._selection_changed, table, sheet_name))
        for i in range(len(list_for_x)):
            table.setColumnWidth(i, 30)
        horizontal_header = ['%0.2f' % x for x in list_for_x]

        table.setHorizontalHeaderLabels(horizontal_header)
        vertical_header = []
        for key in new_list_of_cable:
            vertical_header.append(str(key))
        table.setVerticalHeaderLabels(vertical_header)
        # self._table.itemChanged.connect(self._table_item_is_changed)
        new_matrix_of_yi = []
        for row, key in enumerate(new_list_of_cable):
            new_row = self._add_a_cable_to_table(cable=new_list_of_cable[key],
                                                 row=row, table=table,
                                                 n=len(list_for_x))
            new_matrix_of_yi.append(new_row)
            print(new_matrix_of_yi)

        self._dict_matrix_of_yi[sheet_name] = new_matrix_of_yi

    def get_a_new_list_of_cable(self, sheet_name: str) -> dict:
        return self.dict_of_cables[sheet_name]

    @staticmethod
    def _add_a_cable_to_table(cable: Cable, row: int, table: QTableWidget, n: int) -> []:
        new_row = [0] * n
        for i in range(cable.i_0, cable.i_n + 1):
            table.setItem(row, i, QTableWidgetItem(str(cable.coordinate[i - cable.i_0][1])))
            new_row[i] = cable.coordinate[i - cable.i_0][1]
        return new_row

    def _selection_changed(self, table, sheet_name: str):
        list_of_selected_items = table.selectedIndexes()
        rows = [x.row() for x in list_of_selected_items]
        list_of_hash_ids = list(self.dict_of_cables[sheet_name])
        for name_of_sheet, group_of_cables in self.dict_of_cables.items():
            for key, cable in group_of_cables.items():
                cable.selected = False
        for row in rows:
            cable: Cable = self.dict_of_cables[sheet_name][list_of_hash_ids[row]]
            cable.selected = True

        self._draw_graph()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = WindowSection()
    window.show()
    app.exec()
